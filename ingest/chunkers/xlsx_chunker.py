import hashlib
import os
from typing import Iterator

import openpyxl
import tiktoken

from .base import Chunk

CHUNK_SIZE = 800
CHUNK_OVERLAP = 150
ENCODER = tiktoken.get_encoding("cl100k_base")


def _token_count(text: str) -> int:
    return len(ENCODER.encode(text))


def _row_to_cells(cells) -> list[str]:
    def _clean(v) -> str:
        if v is None:
            return ""
        # Newlines (Alt+Entrée dans Excel) et pipes cassent les tables Markdown
        return str(v).replace("\r\n", " ").replace("\n", " ").replace("\r", " ").replace("|", "\\|").strip()
    return [_clean(c.value) for c in cells]


def _is_empty_row(cells: list[str]) -> bool:
    return not any(v for v in cells)


def _fmt_row(cells: list[str], width: int) -> str:
    padded = [cells[i] if i < len(cells) else "" for i in range(width)]
    return "| " + " | ".join(padded) + " |"


def _build_md_table(headers: list[str], rows: list[list[str]]) -> str:
    w = len(headers)
    separator = "| " + " | ".join("---" for _ in headers) + " |"
    lines = [_fmt_row(headers, w), separator] + [_fmt_row(r, w) for r in rows]
    return "\n".join(lines)


class XLSXChunker:
    def chunk_file(self, file_path: str) -> Iterator[Chunk]:
        with open(file_path, "rb") as f:
            raw_bytes = f.read()

        file_hash = hashlib.sha256(raw_bytes).hexdigest()
        source_file = os.path.basename(file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        chunk_index = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows())
            if not rows:
                continue

            raw_headers = _row_to_cells(rows[0])
            has_headers = any(h and not h.startswith("None") for h in raw_headers)
            headers = raw_headers if has_headers else [f"Col{i+1}" for i in range(len(raw_headers))]
            data_rows = rows[1:] if has_headers else rows

            # Coût fixe du header (répété à chaque chunk pour qu'il soit autonome)
            header_overhead = _token_count(
                _fmt_row(headers, len(headers)) + "\n" +
                "| " + " | ".join("---" for _ in headers) + " |"
            )

            buffer: list[list[str]] = []
            buffer_tokens = header_overhead

            def _emit(buf: list[list[str]]) -> Chunk:
                nonlocal chunk_index
                c = Chunk(
                    content=_build_md_table(headers, buf),
                    source_file=source_file,
                    page_number=1,
                    chunk_index=chunk_index,
                    doc_type="xlsx",
                    section=sheet_name,
                    title=source_file,
                    file_hash=file_hash,
                )
                chunk_index += 1
                return c

            for row in data_rows:
                cells = _row_to_cells(row)
                if _is_empty_row(cells):
                    continue

                row_tokens = _token_count(_fmt_row(cells, len(headers)))

                if buffer_tokens + row_tokens > CHUNK_SIZE and buffer:
                    yield _emit(buffer)

                    # Overlap : dernières lignes jusqu'à CHUNK_OVERLAP tokens
                    overlap: list[list[str]] = []
                    ot = 0
                    for prev in reversed(buffer):
                        lt = _token_count(_fmt_row(prev, len(headers)))
                        if ot + lt <= CHUNK_OVERLAP:
                            overlap.insert(0, prev)
                            ot += lt
                        else:
                            break
                    buffer = overlap
                    buffer_tokens = header_overhead + ot

                buffer.append(cells)
                buffer_tokens += row_tokens

            if buffer:
                yield _emit(buffer)

        wb.close()
